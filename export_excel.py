#!/usr/bin/env python3
"""Export match data to Excel spreadsheet with tabs by decade."""

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from storage.database import Database


def create_excel_by_decade(db_path: str = "data/matches.db", output_path: str = "data/matches_by_decade.xlsx"):
    """Create Excel spreadsheet with matches grouped by decade."""
    db = Database(db_path)
    matches = db.get_all_matches()

    # Group matches by decade
    decades = defaultdict(list)
    for match in matches:
        decade = (match.date.year // 10) * 10
        decades[decade].append(match)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Date", "Opposition", "Venue", "Goals For", "Goals Against", "Result", "Competition", "Season"]
    col_widths = [12, 25, 8, 10, 12, 8, 20, 12]

    # Create a sheet for each decade
    for decade in sorted(decades.keys()):
        sheet_name = f"{decade}s"
        ws = wb.create_sheet(title=sheet_name)

        # Add headers
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Add match data
        for row, match in enumerate(sorted(decades[decade], key=lambda x: x.date), 2):
            ws.cell(row=row, column=1, value=match.date.strftime("%Y-%m-%d")).border = thin_border
            ws.cell(row=row, column=2, value=match.opposition).border = thin_border
            ws.cell(row=row, column=3, value=match.venue).border = thin_border
            ws.cell(row=row, column=4, value=match.goals_for).border = thin_border
            ws.cell(row=row, column=5, value=match.goals_against).border = thin_border
            ws.cell(row=row, column=6, value=match.result).border = thin_border
            ws.cell(row=row, column=7, value=match.competition).border = thin_border
            ws.cell(row=row, column=8, value=match.season).border = thin_border

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return decades


def main():
    output_path = "data/matches_by_decade.xlsx"

    if len(sys.argv) > 1:
        output_path = sys.argv[1]

    decades = create_excel_by_decade(output_path=output_path)

    total = sum(len(matches) for matches in decades.values())
    print(f"Created {output_path} with {len(decades)} decade tabs ({total} matches)")
    for decade in sorted(decades.keys()):
        print(f"  {decade}s: {len(decades[decade])} matches")


if __name__ == "__main__":
    main()
